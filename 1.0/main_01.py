# main.py
import os
import glob
import zipfile
import cv2
import pytesseract
import openpyxl
import time
import tkinter as tk

from datetime import datetime
from tkinter import font
from PIL import Image, ImageTk
from winotify import Notification, audio
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service

from interface import criar_interface
from processamento import iniciar_processamento
from forms import abrir_forms_e_preencher
import openpyxl
from datetime import datetime

def iniciar_programa():
    #root.destroy()  # Fecha a janela inicial
    iniciar_processamento(nomes_moradores, abrir_forms_e_preencher)


def criar_planilha_entrega(nome_planilha):
    if not os.path.exists(nome_planilha):
        planilha = openpyxl.Workbook()
        folha = planilha.active
        folha.append(["N° Entrega", "Nome", "Data", "Horário"])
        planilha.save(nome_planilha)

def registrar_entrega(nome_planilha, nome, data, horario):
    planilha = openpyxl.load_workbook(nome_planilha)
    folha = planilha.active

    num_entrega = 1
    for row in folha.iter_rows(min_row=2, max_row=folha.max_row, values_only=True):
        if row[1] == nome and row[2] == data:
            num_entrega += 1
    
    folha.append([num_entrega, nome, data, horario])
    planilha.save(nome_planilha)

def exibir_notificacao_sucesso(nome_pessoa):
    title = "Entrega Registrada"
    message = f"A encomenda de {nome_pessoa} foi registrada com sucesso!"
    Notification(title, message).show()
    time.sleep(5)


def compactar_planilhas_mes_anterior():
    hoje = datetime.now()
    mes_anterior = hoje.month - 1 if hoje.month > 1 else 12
    ano_anterior = hoje.year if hoje.month > 1 else hoje.year - 1
    pasta_mes_anterior = f"{ano_anterior}-{mes_anterior:02d}"

    if not os.path.exists(pasta_mes_anterior):
        return  

    arquivos_xlsx = glob.glob(os.path.join(pasta_mes_anterior, "*.xlsx"))
    if arquivos_xlsx:
        nome_zip = f"Planilhas_{pasta_mes_anterior}.zip"
        with zipfile.ZipFile(nome_zip, "w") as zipf:
            for arquivo in arquivos_xlsx:
                zipf.write(arquivo, os.path.basename(arquivo))
        
        for arquivo in arquivos_xlsx:
            os.remove(arquivo)
        os.rmdir(pasta_mes_anterior)

planilha_moradores = openpyxl.load_workbook('PLANILHA.xlsx')
folha_moradores = planilha_moradores.active

nomes_moradores = {}
for linha in folha_moradores.iter_rows(values_only=True):
    nome = linha[0].lower()
    nomes_moradores[nome] = {"apartamento": linha[1], "bloco": linha[2]}

nome_planilha_entrega = f"Entregas_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
criar_planilha_entrega(nome_planilha_entrega)

criar_interface(iniciar_programa)
#cocolascado