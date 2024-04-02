import os
import glob
import zipfile
import cv2
import pytesseract
import openpyxl
import time
import tkinter as tk

from datetime import datetime
from tkinter import font
from PIL import Image, ImageTk
from winotify import Notification, audio
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service


import tkinter as tk
from tkinter import font

def iniciar_programa():
    root.destroy()  # Fecha a janela inicial
    # Aqui vai o restante do seu código

root = tk.Tk()
root.title("Tela Inicial")
root.geometry("600x300")
root.configure(bg="#F0F0F0")  # Cor de fundo
root.resizable(False, False)  # Impede a redimensionamento da janela

# Define uma fonte personalizada
fonte_titulo = font.Font(family="Helvetica", size=16, weight="bold")
fonte_texto = font.Font(family="Helvetica", size=12)

# Carrega a imagem do ícone da caixa
icone_caixa = Image.open("caixa.png")
icone_caixa = icone_caixa.resize((100, 100))  # Redimensiona a imagem
icone_caixa = ImageTk.PhotoImage(icone_caixa)

# Adiciona o ícone da caixa como um rótulo
label_icone = tk.Label(root, image=icone_caixa, bg="#F0F0F0")
label_icone.pack(pady=10)

# Adiciona um rótulo de boas-vindas
label = tk.Label(root, text="Bem-vindo!", font=fonte_titulo, bg="#F0F0F0", fg="#333333")
label.pack()

# Adiciona um rótulo com instruções
instrucoes = tk.Label(root, text="Clique em 'Iniciar' para começar.", font=fonte_texto, bg="#F0F0F0", fg="#666666")
instrucoes.pack()

# Adiciona um botão "Iniciar"
button = tk.Button(root, text="Iniciar", command=iniciar_programa, bg="#4CAF50", fg="white", relief="flat")
button.config(width=10, height=2, font=fonte_texto)
button.pack(pady=20)

root.mainloop()



# Função para criar uma nova planilha Excel com as informações da entrega
def criar_planilha_entrega(nome_planilha):
    if not os.path.exists(nome_planilha):
        planilha = openpyxl.Workbook()
        folha = planilha.active
        folha.append(["N° Entrega", "Nome", "Data", "Horário"])
        planilha.save(nome_planilha)

# Função para registrar uma entrega na planilha Excel
def registrar_entrega(nome_planilha, nome, data, horario):
    planilha = openpyxl.load_workbook(nome_planilha)
    folha = planilha.active

    # Obter o número da próxima entrega para o mesmo morador na mesma data
    num_entrega = 1
    for row in folha.iter_rows(min_row=2, max_row=folha.max_row, values_only=True):
        if row[1] == nome and row[2] == data:
            num_entrega += 1
    
    # Adicionar a nova entrega à planilha
    folha.append([num_entrega, nome, data, horario])
    planilha.save(nome_planilha)


# Função para exibir uma notificação de sucesso usando winotify
def exibir_notificacao_sucesso(nome_pessoa):
    title = "Entrega Registrada"
    message = f"A encomenda de {nome_pessoa} foi registrada com sucesso!"
    Notification(title, message).show()
    time.sleep(5)  # Aguarde 5 segundos para a notificação ser exibida

    # Função para compactar planilhas do mês anterior em um arquivo zip
def compactar_planilhas_mes_anterior():
    hoje = datetime.now()
    mes_anterior = hoje.month - 1 if hoje.month > 1 else 12
    ano_anterior = hoje.year if hoje.month > 1 else hoje.year - 1
    pasta_mes_anterior = f"{ano_anterior}-{mes_anterior:02d}"

    if not os.path.exists(pasta_mes_anterior):
        return  # Não há planilhas do mês anterior para compactar

    arquivos_xlsx = glob.glob(os.path.join(pasta_mes_anterior, "*.xlsx"))
    if arquivos_xlsx:
        nome_zip = f"Planilhas_{pasta_mes_anterior}.zip"
        with zipfile.ZipFile(nome_zip, "w") as zipf:
            for arquivo in arquivos_xlsx:
                zipf.write(arquivo, os.path.basename(arquivo))
        # Remover as planilhas originais após compactação
        for arquivo in arquivos_xlsx:
            os.remove(arquivo)
        # Remover o diretório vazio do mês anterior
        os.rmdir(pasta_mes_anterior)
# Função para pré-processar a imagem antes do OCR
def preprocess_image(image):
    # Convertendo para escala de cinza
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Aplicando binarização
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)
    return binary




# Carregar a planilha do Excel com os nomes dos moradores
planilha_moradores = openpyxl.load_workbook('PLANILHA.xlsx')
folha_moradores = planilha_moradores.active

# Criar um dicionário para armazenar os nomes dos moradores
nomes_moradores = {}
for linha in folha_moradores.iter_rows(values_only=True):
    nome = linha[0].lower()  # Supondo que o nome esteja na primeira coluna e transformando para minúsculas
    nomes_moradores[nome] = {"apartamento": linha[1], "bloco": linha[2]}

# Inicializar a câmera
camera = cv2.VideoCapture(0)

# Criar uma nova planilha para o dia atual
data_atual = datetime.now().strftime("%Y-%m-%d")
nome_planilha_entrega = f"Entregas_{data_atual}.xlsx"
criar_planilha_entrega(nome_planilha_entrega)

while True:
    # Capturar o próximo frame da câmera
    validacao, frame = camera.read()

    # Verificar se o frame foi capturado corretamente
    if not validacao:
        print("Erro ao capturar o próximo frame.")
        break

    # Exibir o frame na janela
    cv2.imshow("Video", frame)
# Pré-processamento da imagem
    preprocessed_frame = preprocess_image(frame)
    # Converter o frame para escala de cinza para melhorar o desempenho do OCR
   # frame_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Realizar OCR para extrair o texto do frame
    texto_extraido = pytesseract.image_to_string(preprocessed_frame)

    # Comparar os nomes extraídos com os nomes dos moradores na planilha
    for nome, info in nomes_moradores.items():
        if nome in texto_extraido.lower():  # Converter para minúsculas para garantir a comparação
            print("Nome do destinatário encontrado:", nome)
            print("Apartamento:", info["apartamento"])
            print("Bloco:", info["bloco"])

           # Abrir o formulário no navegador
            url_do_forms = "https://forms.gle/av9HDTPS1FBEmPJZ7"
            service = Service(executable_path='chromedriver.exe')
            driver = webdriver.Chrome(service=service)
            driver.get(url_do_forms)
            time.sleep(2)


            # Preencher as informações do formulário com base na correspondência encontrada
            elemento_texto_nome = driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div/div[1]/input')
            elemento_apartamento = driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div[1]/input')
            elemento_bloco = driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div/div[1]/input')

            elemento_texto_nome.send_keys(nome)
            elemento_apartamento.send_keys(info["apartamento"])
            elemento_bloco.send_keys(info["bloco"])

            time.sleep(2)
            driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div[1]/div/span').click()

            # Fechar o navegador após o preenchimento do formulário
            driver.quit()  

            # Registrar a entrega na planilha Excel
            data_hora_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            registrar_entrega(nome_planilha_entrega, nome, data_atual, data_hora_atual)

            # Exibir uma mensagem de confirmação em uma janela pop-up
            # mensagem = f"A Encomenda de {nome} foi registrada com Sucesso!"
            exibir_notificacao_sucesso(nome)


    # Verificar se a tecla ESC foi pressionada para sair do loop
    key = cv2.waitKey(1)
    if key == 27:  # ESC
        break

# Liberar a câmera e fechar todas as janelas
camera.release()
cv2.destroyAllWindows()
