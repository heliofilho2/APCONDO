import os
import glob
import zipfile
import cv2
import pytesseract
import openpyxl
import time
from datetime import datetime
import tkinter as tk

from winotify import Notification, audio
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service


# Função para criar uma nova planilha Excel com as informações da entrega
def criar_planilha_entrega(nome_planilha):
    if not os.path.exists(nome_planilha):
        planilha = openpyxl.Workbook()
        folha = planilha.active
        folha.append(["N° Entrega", "Nome", "Data", "Horário"])
        planilha.save(nome_planilha)

# Função para registrar uma entrega na planilha Excel
def registrar_entrega(nome_planilha, nome, data, horario):
    planilha = openpyxl.load_workbook(nome_planilha)
    folha = planilha.active

    # Obter o número da próxima entrega para o mesmo morador na mesma data
    num_entrega = 1
    for row in folha.iter_rows(min_row=2, max_row=folha.max_row, values_only=True):
        if row[1] == nome and row[2] == data:
            num_entrega += 1
    
    # Adicionar a nova entrega à planilha
    folha.append([num_entrega, nome, data, horario])
    planilha.save(nome_planilha)


# Função para exibir uma notificação de sucesso usando winotify
def exibir_notificacao_sucesso(nome_pessoa):
    title = "Entrega Registrada"
    message = f"A encomenda de {nome_pessoa} foi registrada com sucesso!"
    Notification(title, message).show()
    time.sleep(5)  # Aguarde 5 segundos para a notificação ser exibida

# Função para realizar a segunda validação do apartamento caso haja nomes iguais na planilha
def validar_apartamento(nome, info_moradores, texto_extraido):
    apartamento = None
    if nome in info_moradores:
        apartamento = info_moradores[nome]["apartamento"]
        if isinstance(apartamento, str):
            if len(apartamento) > 1:
                # Verifica se o texto extraído contém o número do apartamento
                if apartamento in texto_extraido:
                    return apartamento
                else:
                    apartamento_validado = input(f"Por favor, confirme o número do apartamento para {nome}: ")
                    if apartamento_validado == apartamento:
                        return apartamento
                    else:
                        print("Número do apartamento não corresponde ao registrado na planilha.")
                        return None
            else:
                return apartamento
    return apartamento

# Função para criar e exibir a interface gráfica
def exibir_interface_grafica():
    root = tk.Tk()
    root.title("Entrega de Encomendas")

    # Função para registrar a entrega quando o botão "Registrar Entrega" for clicado
    def registrar_entrega():
        nome = entry_nome.get()
        apartamento = entry_apartamento.get()

        # Verificar se o nome e o número do apartamento foram fornecidos
        if nome and apartamento:
            # Registrar a entrega na planilha
            data_atual = datetime.now().strftime("%Y-%m-%d")
            hora_atual = datetime.now().strftime("%H:%M:%S")
            registrar_entrega("Entregas.xlsx", nome, data_atual, hora_atual)

            # Exibir notificação de sucesso
            exibir_notificacao_sucesso(nome)
        else:
            # Exibir uma mensagem de erro se o nome ou o número do apartamento estiverem em falta
            messagebox.showerror("Erro", "Por favor, forneça o nome e o número do apartamento.")

    # Criar e posicionar widgets na interface
    label_nome = tk.Label(root, text="Nome:")
    label_nome.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)

    entry_nome = tk.Entry(root)
    entry_nome.grid(row=0, column=1, padx=10, pady=5)

    label_apartamento = tk.Label(root, text="Apartamento:")
    label_apartamento.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)

    entry_apartamento = tk.Entry(root)
    entry_apartamento.grid(row=1, column=1, padx=10, pady=5)

    button_registrar = tk.Button(root, text="Registrar Entrega", command=registrar_entrega)
    button_registrar.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

    root.mainloop()

# Carregar a planilha do Excel com os nomes dos moradores
planilha_moradores = openpyxl.load_workbook('PLANILHA.xlsx')
folha_moradores = planilha_moradores.active

# Criar um dicionário para armazenar os nomes dos moradores
nomes_moradores = {}
for linha in folha_moradores.iter_rows(values_only=True):
    nome = linha[0].lower()  # Supondo que o nome esteja na primeira coluna e transformando para minúsculas
    nomes_moradores[nome] = {"apartamento": linha[1], "bloco": linha[2]}

# Inicializar a câmera
camera = cv2.VideoCapture(0)

# Criar uma nova planilha para o dia atual
data_atual = datetime.now().strftime("%Y-%m-%d")
nome_planilha_entrega = f"Entregas_{data_atual}.xlsx"
criar_planilha_entrega(nome_planilha_entrega)

# Exibir a interface gráfica
exibir_interface_grafica()

while True:
    # Capturar o próximo frame da câmera
    validacao, frame = camera.read()

    # Verificar se o frame foi capturado corretamente
    if not validacao:
        print("Erro ao capturar o próximo frame.")
        break

    # Exibir o frame na janela
    cv2.imshow("Video", frame)

    # Converter o frame para escala de cinza para melhorar o desempenho do OCR
    frame_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Realizar OCR para extrair o texto do frame
    texto_extraido = pytesseract.image_to_string(frame_gray)

    # Comparar os nomes extraídos com os nomes dos moradores na planilha
    for nome, info in nomes_moradores.items():
        if nome in texto_extraido.lower():  # Converter para minúsculas para garantir a comparação
            print("Nome do destinatário encontrado:", nome)
            print("Apartamento:", info["apartamento"])
            print("Bloco:", info["bloco"])

            apartamento = validar_apartamento(nome, nomes_moradores, texto_extraido)
            if apartamento:
                # Registrar a entrega na planilha Excel
                data_hora_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                registrar_entrega(nome_planilha_entrega, nome, data_atual, data_hora_atual)

                # Exibir uma mensagem de confirmação em uma janela pop-up
                exibir_notificacao_sucesso(nome)

    # Verificar se a tecla ESC foi pressionada para sair do loop
    key = cv2.waitKey(1)
    if key == 27:  # ESC
        break

# Liberar a câmera e fechar todas as janelas
camera.release()
cv2.destroyAllWindows()
